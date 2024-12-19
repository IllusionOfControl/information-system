from datetime import timedelta, datetime

import openpyxl
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import F, Value
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect

from information_system.apps.core.forms import RepairForm, MaintenanceForm
from information_system.apps.core.models import Department, DeviceModel, DeviceInstance, DeviceType, Repair, \
    Maintenance, MalfunctionType, User, MaintenanceType


def index(request):
    return render(request, 'core/index.html')


@login_required
def profile(request):
    user = request.user

    recent_repairs = Repair.objects.filter(performed_by=user).order_by('-date_repair')[:5]

    recent_maintenance = Maintenance.objects.filter(performed_by=user).order_by('-date')[:5]

    total_repairs = Repair.objects.filter(performed_by=user).count()
    total_maintenance = Maintenance.objects.filter(performed_by=user).count()

    last_month = datetime.now() - timedelta(days=30)
    monthly_repairs = Repair.objects.filter(performed_by=user, date_repair__gte=last_month).count()
    monthly_maintenance = Maintenance.objects.filter(performed_by=user, date__gte=last_month).count()

    context = {
        'user': user,
        'recent_repairs': recent_repairs,
        'recent_maintenance': recent_maintenance,
        'total_repairs': total_repairs,
        'total_maintenance': total_maintenance,
        'monthly_repairs': monthly_repairs,
        'monthly_maintenance': monthly_maintenance,
    }

    return render(request, 'core/profile.html', context)


def department_list(request):
    departments = Department.objects.all()
    selected_department = None
    devices_in_department = []

    if request.GET.get('department'):
        selected_department = get_object_or_404(Department, pk=request.GET.get('department'))
        devices_in_department = DeviceInstance.objects.filter(department=selected_department).select_related(
            'device_model')

    context = {
        'departments': departments,
        'selected_department': selected_department,
        'devices_in_department': devices_in_department,
    }
    return render(request, 'core/department_list.html', context)


def device_list(request):
    devices = DeviceModel.objects.all()
    context = {'devices': devices}
    return render(request, 'core/device_list.html', context)


def device_model_list(request):
    device_models = DeviceModel.objects.all().select_related('type')
    device_types = DeviceType.objects.all()
    selected_device_model = None
    device_instances = []

    if request.GET.get('device_type'):
        selected_device_type = get_object_or_404(DeviceType, pk=request.GET.get('device_type'))
        device_instances = DeviceInstance.objects.filter(device_model__type=selected_device_type).select_related(
            'device_model')

    if request.GET.get('device_model'):
        selected_device_model = get_object_or_404(DeviceModel, pk=request.GET.get('device_model'))
        device_instances = DeviceInstance.objects.filter(device_model=selected_device_model)

    context = {
        'device_models': device_models,
        'device_types': device_types,
        'selected_device_model': selected_device_model,
        'device_instances': device_instances,
    }
    return render(request, 'core/device_model_list.html', context)


def repair_list(request):
    queryset = Repair.objects.all().select_related(
        'device_instance', 'device_instance__device_model', 'performed_by', "malfunction_type"
    )

    device_name = request.GET.get('device_name')
    inventory_number = request.GET.get('inventory_number')
    date_breakdown_start = request.GET.get('date_breakdown_start')
    date_breakdown_end = request.GET.get('date_breakdown_end')
    device_type = request.GET.get('device_type')
    malfunction_type = request.GET.get('malfunction_type')
    performed_by = request.GET.get('performed_by')
    department = request.GET.get('department')

    if device_name:
        queryset = (
            queryset
            .annotate(full_name=Concat(
                F('device_instance__device_model__manufacturer'),
                Value(' '),
                F('device_instance__device_model__name')
            ))
            .filter(full_name__icontains=device_name)
        )

    if inventory_number:
        queryset = queryset.filter(device_instance__inventory_number=inventory_number)

    if date_breakdown_start:
        queryset = queryset.filter(date_breakdown__gte=date_breakdown_start)

    if date_breakdown_end:
        queryset = queryset.filter(date_breakdown__lte=date_breakdown_end)

    if device_type:
        queryset = queryset.filter(device_instance__device_model__type_id=device_type)

    if malfunction_type:
        queryset = queryset.filter(malfunction_type_id=malfunction_type)

    if performed_by:
        queryset = queryset.filter(performed_by_id=performed_by)

    if department:
        queryset = queryset.filter(device_instance__department_id=department)

    repairs = queryset

    if "submit" in request.GET:
        device_types = DeviceType.objects.all()
        malfunction_types = MalfunctionType.objects.all()
        users = User.objects.all()
        departments = Department.objects.all()

        context = {'repairs': repairs, 'device_types': device_types, 'malfunction_types': malfunction_types,
                   'users': users,
                   'departments': departments}
        return render(request, 'core/repair_list.html', context)
    if "export" in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="repairs.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Ремонты"

        # Заголовки столбцов
        columns = [
            "Дата поломки", "Устройство", "Тип неисправности", "Дата устранения", "Выполнил", "Примечание"
        ]
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Данные
        for row_num, repair in enumerate(queryset, 2):  # Начинаем со второй строки
            worksheet.cell(row=row_num, column=1).value = repair.date_breakdown.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=2).value = str(repair.device_instance)
            worksheet.cell(row=row_num, column=3).value = str(repair.malfunction_type)
            worksheet.cell(row=row_num, column=4).value = repair.date_repair.strftime(
                '%Y-%m-%d') if repair.date_repair else ""  # Проверка на None
            worksheet.cell(row=row_num, column=5).value = str(repair.performed_by.FIO)
            worksheet.cell(row=row_num, column=6).value = repair.note

        workbook.save(response)

        return response
    else:
        device_types = DeviceType.objects.all()
        malfunction_types = MalfunctionType.objects.all()
        users = User.objects.all()
        departments = Department.objects.all()

        context = {'repairs': repairs, 'device_types': device_types, 'malfunction_types': malfunction_types,
                   'users': users,
                   'departments': departments}
        return render(request, 'core/repair_list.html', context)


def maintenance_list(request):
    queryset = Maintenance.objects.all().select_related('device_instance', 'device_instance__device_model',
                                                        'performed_by', 'type')

    device_name = request.GET.get('device_name')
    inventory_number = request.GET.get('inventory_number')
    date_breakdown_start = request.GET.get('date_breakdown_start')
    date_breakdown_end = request.GET.get('date_breakdown_end')
    device_type = request.GET.get('device_type')
    malfunction_type = request.GET.get('malfunction_type')
    performed_by = request.GET.get('performed_by')
    department = request.GET.get('department')

    if inventory_number:
        queryset = queryset.filter(device_instance__inventory_number=inventory_number)

    if device_name:
        queryset = (
            queryset
            .annotate(full_name=Concat(
                F('device_instance__device_model__manufacturer'),
                Value(' '),
                F('device_instance__device_model__name')
            ))
            .filter(full_name__icontains=device_name)
        )

    if inventory_number:
        queryset = queryset.filter(device_instance__inventory_number=inventory_number)

    if date_breakdown_start:
        queryset = queryset.filter(date__gte=date_breakdown_start)

    if date_breakdown_end:
        queryset = queryset.filter(date__lte=date_breakdown_end)

    if device_type:
        queryset = queryset.filter(device_instance__device_model__type_id=device_type)

    if malfunction_type:
        queryset = queryset.filter(type__name=malfunction_type)

    if performed_by:
        queryset = queryset.filter(performed_by_id=performed_by)

    if department:
        queryset = queryset.filter(device_instance__department_id=department)

    maintenances = queryset
    device_types = DeviceType.objects.all()
    maintenance_types = MaintenanceType.objects.all()
    users = User.objects.all()
    device_instances = DeviceInstance.objects.all()
    departments = Department.objects.all()

    if 'submit' in request.GET:
        context = {
            'maintenances': maintenances,
            'device_types': device_types,
            'maintenance_types': maintenance_types,
            'users': users,
            'device_instances': device_instances,
            'departments': departments,
        }
        return render(request, 'core/maintenance_list.html', context)
    if "export" in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="maintenance.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Техническое обслуживание"

        # Заголовки столбцов
        columns = [
            "Дата ТО", "Отделение ОПС", "Инвентарный номер", "Тип устройства", "Модель Устройства", "Тип обслуживания",
            "Выполнил",
            "Примечание"
        ]
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Данные
        for row_num, maintenance in enumerate(queryset, 2):  # Начинаем со второй строки
            worksheet.cell(row=row_num, column=1).value = maintenance.date.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=2).value = str(maintenance.device_instance.department.name)
            worksheet.cell(row=row_num, column=3).value = str(maintenance.device_instance.inventory_number)
            worksheet.cell(row=row_num, column=4).value = str(maintenance.device_instance.device_model.type)
            worksheet.cell(row=row_num, column=5).value = str(maintenance.device_instance.device_model.name)
            worksheet.cell(row=row_num, column=6).value = str(maintenance.type.name)
            worksheet.cell(row=row_num, column=7).value = str(maintenance.performed_by.FIO)
            worksheet.cell(row=row_num, column=8).value = maintenance.note

        workbook.save(response)

        return response
    else:
        context = {
            'maintenances': maintenances,
            'device_types': device_types,
            'maintenance_types': maintenance_types,
            'users': users,
            'device_instances': device_instances,
            'departments': departments,
        }
        return render(request, 'core/maintenance_list.html', context)


def master_required(view_func):
    decorated_view_func = login_required(user_passes_test(lambda u: u.role in ['master', 'administrator'])(view_func))
    return decorated_view_func


@master_required
def create_repair(request):
    if request.method == 'POST':
        form = RepairForm(request.POST, user=request.user)
        if form.is_valid():
            repair = form.save()
            return redirect('repair_list')
    else:
        form = RepairForm(user=request.user)
    return render(request, 'core/create_repair.html', {'form': form})


@master_required
def edit_repair(request, repair_id):
    repair = get_object_or_404(Repair, pk=repair_id)
    if request.method == 'POST':
        form = RepairForm(request.POST, user=request.user, instance=repair)
        if form.is_valid():
            form.save()
            return redirect('repair_list')
    else:
        form = RepairForm(user=request.user, instance=repair)
    return render(request, 'core/edit_repair.html', {'form': form, 'repair': repair})


@master_required
def delete_repair(request, repair_id):
    repair = get_object_or_404(Repair, pk=repair_id)
    repair.delete()
    return redirect('repair_list')  # Перенаправляем на страницу списка ремонтов


@master_required
def create_maintenance(request):
    if request.method == 'POST':
        form = MaintenanceForm(request.POST, user=request.user)
        if form.is_valid():
            maintenance = form.save()
            return redirect('maintenance_list')
    else:
        form = MaintenanceForm(user=request.user)
    return render(request, 'core/create_maintenance.html', {'form': form})


@master_required
def edit_maintenance(request, maintenance_id):
    maintenance = get_object_or_404(Maintenance, pk=maintenance_id)
    if request.method == 'POST':
        form = MaintenanceForm(request.POST, user=request.user, instance=maintenance)
        if form.is_valid():
            form.save()
            return redirect('maintenance_list')
    else:
        form = MaintenanceForm(user=request.user, instance=maintenance)
    return render(request, 'core/edit_maintenance.html', {'form': form, 'maintenance': maintenance})


@master_required
def delete_maintenance(request, maintenance_id):
    repair = get_object_or_404(Maintenance, pk=maintenance_id)
    repair.delete()
    return redirect('maintenance_list')


from django.shortcuts import render
import requests


def exchange_rates(request):
    url = 'https://www.nbrb.by/api/exrates/rates?periodicity=0'  # API Нацбанка Беларуси для ежедневных курсов
    error_message = None
    try:
        response = requests.get(url)
        response.raise_for_status()  # Проверка на ошибки HTTP
        rates_data = response.json()
    except requests.exceptions.RequestException as e:
        rates_data = []  # Пустой список, если ошибка
        error_message = f"Ошибка получения данных о курсах валют: {e}"

    amount = request.GET.get('amount', '')
    from_currency = request.GET.get('from_currency', '')
    to_currency = request.GET.get('to_currency', '')
    converted_amount = None
    error_conversion = None

    if amount and from_currency and to_currency:
        try:
            amount = float(amount)
            from_rate = next((rate['Cur_OfficialRate'] / rate['Cur_Scale'] for rate in rates_data if
                              rate['Cur_Abbreviation'] == from_currency), None)
            to_rate = next((rate['Cur_OfficialRate'] / rate['Cur_Scale'] for rate in rates_data if
                            rate['Cur_Abbreviation'] == to_currency), None)
            if from_rate and to_rate:
                converted_amount = amount * from_rate / to_rate
            else:
                error_conversion = "Выбраны неверные валюты."

        except ValueError:
            error_conversion = "Неверный формат суммы."

    context = {
        'rates_data': rates_data,
        'amount': amount,
        'from_currency': from_currency,
        'to_currency': to_currency,
        'converted_amount': converted_amount,
        'error_conversion': error_conversion,
        'error_message': error_message,
    }
    return render(request, 'core/exchange_rates.html', context)
